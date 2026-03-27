"""
여러 파일을 통합하는 유틸리티
"""

import pandas as pd
from typing import List, Optional
import os
import logging
import utils
from preprocessor.apc_preprocessor import flatten_multilevel_header


class DataMerger:
    """데이터 파일 통합 클래스"""

    def __init__(self, logger: Optional[logging.Logger] = None):
        """
        Parameters:
        -----------
        logger : logging.Logger
            로거 객체
        """
        self.logger = logger or logging.getLogger('coating_preprocessor.merger')

    def merge_apc_files(self, file_list: List[str]) -> Optional[pd.DataFrame]:
        """
        여러 APC 파일을 하나로 통합

        Parameters:
        -----------
        file_list : List[str]
            통합할 파일 경로 리스트

        Returns:
        --------
        pd.DataFrame
            통합된 데이터프레임
        """
        self.logger.info("="*80)
        self.logger.info("APC 파일 통합 시작")
        self.logger.info("="*80)

        all_data = []

        for file_path in file_list:
            self.logger.info(f"처리 중: {os.path.basename(file_path)}")

            try:
                # 파일 형식에 따라 자동 로드 (Excel: xlwings/openpyxl, CSV/Parquet: pandas)
                df = utils.load_file(file_path, logger=self.logger)

                self.logger.info(f"  ✓ {len(df)} 행 로드")
                all_data.append(df)

            except Exception as e:
                self.logger.error(f"  ✗ 오류: {e}")
                continue

        if not all_data:
            self.logger.error("통합할 데이터가 없습니다.")
            return None

        # 데이터 통합
        merged_df = pd.concat(all_data, ignore_index=True)

        # TIME 칼럼으로 정렬
        if 'TIME' in merged_df.columns:
            merged_df['TIME'] = pd.to_datetime(merged_df['TIME'])
            merged_df = merged_df.sort_values('TIME').reset_index(drop=True)

        self.logger.info("="*80)
        self.logger.info(f"통합 완료: {len(merged_df)} 행")
        self.logger.info("="*80)

        return merged_df

    def merge_densitometer_files(self, file_list: List[str]) -> Optional[pd.DataFrame]:
        """
        여러 밀도계 파일을 하나로 통합

        Parameters:
        -----------
        file_list : List[str]
            통합할 파일 경로 리스트

        Returns:
        --------
        pd.DataFrame
            통합된 데이터프레임
        """
        self.logger.info("="*80)
        self.logger.info("밀도계 파일 통합 시작")
        self.logger.info("="*80)

        all_data = []

        for file_path in file_list:
            self.logger.info(f"처리 중: {os.path.basename(file_path)}")

            try:
                # 파일 형식에 따라 자동 로드 (Excel: xlwings/openpyxl, CSV/Parquet: pandas)
                df = utils.load_file(file_path, logger=self.logger)

                self.logger.info(f"  ✓ {len(df)} 행 로드")
                all_data.append(df)

            except Exception as e:
                self.logger.error(f"  ✗ 오류: {e}")
                continue

        if not all_data:
            self.logger.error("통합할 데이터가 없습니다.")
            return None

        # 데이터 통합
        merged_df = pd.concat(all_data, ignore_index=True)

        # 첫 번째 칼럼(시간)으로 정렬
        time_col = merged_df.columns[0]
        merged_df[time_col] = pd.to_datetime(merged_df[time_col])
        merged_df = merged_df.sort_values(time_col).reset_index(drop=True)

        self.logger.info("="*80)
        self.logger.info(f"통합 완료: {len(merged_df)} 행")
        self.logger.info("="*80)

        return merged_df

    def merge_llspec_files(self, file_list: List[str]) -> Optional[pd.DataFrame]:
        """
        여러 LLspec 파일을 하나로 통합

        LLspec 엑셀 파일은 듀얼 헤더(멀티레벨 헤더)로 구성될 수 있으므로
        flatten_multilevel_header를 사용하여 평탄화 처리

        Parameters:
        -----------
        file_list : List[str]
            통합할 파일 경로 리스트

        Returns:
        --------
        pd.DataFrame
            통합된 데이터프레임
        """
        self.logger.info("="*80)
        self.logger.info("LLspec 파일 통합 시작")
        self.logger.info("="*80)

        all_data = []

        for file_path in file_list:
            self.logger.info(f"처리 중: {os.path.basename(file_path)}")

            try:
                ext = os.path.splitext(file_path)[1].lower()

                if ext in ('.xlsx', '.xls'):
                    # LLspec 엑셀은 듀얼 헤더 가능 → raw 데이터로 읽어서 flatten 처리
                    df = self._load_llspec_excel(file_path)
                else:
                    # Parquet/CSV는 이미 평탄화된 헤더이므로 그대로 로드
                    df = utils.load_file(file_path, logger=self.logger)

                if df is None:
                    continue

                self.logger.info(f"  ✓ {len(df)} 행 로드")
                all_data.append(df)

            except Exception as e:
                self.logger.error(f"  ✗ 오류: {e}")
                continue

        if not all_data:
            self.logger.error("통합할 데이터가 없습니다.")
            return None

        # 데이터 통합
        merged_df = pd.concat(all_data, ignore_index=True)

        # TIME 칼럼으로 정렬
        if 'TIME' in merged_df.columns:
            merged_df['TIME'] = pd.to_datetime(merged_df['TIME'])
            merged_df = merged_df.sort_values('TIME').reset_index(drop=True)

        self.logger.info("="*80)
        self.logger.info(f"통합 완료: {len(merged_df)} 행")
        self.logger.info("="*80)

        return merged_df

    def _load_llspec_excel(self, file_path: str) -> Optional[pd.DataFrame]:
        """
        LLspec 엑셀 파일을 듀얼 헤더 처리하여 로드

        apc_preprocessor의 flatten_multilevel_header를 활용하여
        멀티레벨 헤더를 "대분류_소분류" 형태로 평탄화
        """
        if utils.HAS_XLWINGS:
            import xlwings as xw
            app = xw.App(visible=False)
            wb = None
            try:
                wb = app.books.open(file_path)
                sheet = wb.sheets[0]
                data = sheet.used_range.value
                df = flatten_multilevel_header(data, logger=self.logger)
                if df.empty:
                    self.logger.error(f"  데이터가 비어있습니다: {file_path}")
                    return None
                return df
            except Exception as e:
                self.logger.error(f"  xlwings 로드 오류: {e}")
                return None
            finally:
                if wb:
                    wb.close()
                app.quit()
        else:
            try:
                from openpyxl import load_workbook
                wb = load_workbook(file_path, data_only=True)
                sheet = wb.worksheets[0]
                data = []
                for row in sheet.iter_rows(values_only=True):
                    data.append(list(row))
                wb.close()
                df = flatten_multilevel_header(data, logger=self.logger)
                if df.empty:
                    self.logger.error(f"  데이터가 비어있습니다: {file_path}")
                    return None
                return df
            except Exception as e:
                self.logger.error(f"  openpyxl 로드 오류: {e}")
                return None
